#!/usr/bin/env python3
"""IGT Masraf Takip — Aylık Excel Dışa Aktarım Aracı

Supabase'den belirtilen ayın harcamalarını çekip, şirketin alışık olduğu masraf
formu düzeninde bir .xlsx üretir.

Kullanım:
    # Canlı veriyle (Supabase):
    export SUPABASE_URL="https://....supabase.co"
    export SUPABASE_KEY="service_role_veya_anon_key"
    python igt_excel_export.py --ay 2026-05 --cikti IGT-Masraf-2026-05.xlsx

    # Canlı DB olmadan örnek/deneme çıktısı:
    python igt_excel_export.py --ay 2026-05 --mock --cikti ornek.xlsx
"""
import os
import sys
import argparse
import calendar
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

# ── Ödeme şekli kodu → özet tablo satır etiketi (orijinal formdaki birebir metin)
ODEME_OZET_SATIR = {
    "NAKIT_AVANS": "İGT İş Avansı İle Nakit Harcama",
    "IGT_KK": "İGT Kredi Kartı İle Harcama (YKB Gold Kart)",
    "PERSONEL_NAKIT": "Personel Nakit Harcama",
    "PERSONEL_KK": "Personel Şahsi Kredi Kartı İle Harcama",
}
# Ana tablodaki "H. ŞEKLİ" sütununa yazılacak kısa etiket (SUMIF anahtarı)
ODEME_KISA = {
    "NAKIT_AVANS": "Nakit Avans",
    "IGT_KK": "İGT KK",
    "PERSONEL_NAKIT": "P. Nakit",
    "PERSONEL_KK": "P. Şahsi KK",
}

PARA_FORMAT = "#,##0.00"
TARIH_FORMAT = "DD.MM.YYYY"
FONT_AD = "Arial"

TR_AYLAR = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]


# ────────────────────────────────────────────────────────────────────
# Veri kaynağı
# ────────────────────────────────────────────────────────────────────
def ay_araligi(ay: dt.date):
    ilk = ay.replace(day=1)
    son = ay.replace(day=calendar.monthrange(ay.year, ay.month)[1])
    return ilk.isoformat(), son.isoformat()


def supabase_cek(ay: dt.date):
    """Supabase'den ayın harcamaları, iş avansları ve kategori listesini çeker."""
    from supabase import create_client

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        sys.exit("HATA: SUPABASE_URL ve SUPABASE_KEY ortam değişkenlerini ayarla "
                 "(ya da --mock kullan).")

    sb = create_client(url, key)
    ilk, son = ay_araligi(ay)

    harcamalar = (
        sb.table("harcamalar")
        .select(
            "tarih, firma, fis_no, aciklama, plaka_stok, fis_tutari, kdv,"
            "kategoriler(no, ad),"
            "odeme_sekilleri(kod, ad),"
            "projeler(ad),"
            "personel!harcamalar_personel_id_fkey(ad_soyad)"
        )
        .eq("iptal", False)
        .gte("tarih", ilk)
        .lte("tarih", son)
        .order("tarih")
        .execute()
        .data
    )

    avanslar = (
        sb.table("is_avanslari")
        .select("tarih, tutar, tur, aciklama, personel!is_avanslari_personel_id_fkey(ad_soyad)")
        .eq("iptal", False)
        .gte("tarih", ilk)
        .lte("tarih", son)
        .order("tarih")
        .execute()
        .data
    )

    kategoriler = (
        sb.table("kategoriler")
        .select("no, ad")
        .eq("iptal", False)
        .order("no")
        .execute()
        .data
    )
    return harcamalar, avanslar, kategoriler


def mock_cek(ay: dt.date):
    """Canlı DB olmadan örnek veri — layout doğrulamak için."""
    h = [
        {"tarih": f"{ay:%Y-%m}-01", "firma": "Bolu Zafer Petrol", "fis_no": "27",
         "aciklama": "Araç Yakıt", "plaka_stok": "34GNL417", "fis_tutari": 3500, "kdv": 500,
         "kategoriler": {"no": 4, "ad": "Yakıt"}, "odeme_sekilleri": {"kod": "PERSONEL_KK"},
         "projeler": {"ad": "Bolu Demiryolu"}, "personel": {"ad_soyad": "Halil Karakaş"}},
        {"tarih": f"{ay:%Y-%m}-01", "firma": "A-101", "fis_no": "16",
         "aciklama": "Market", "plaka_stok": None, "fis_tutari": 1500, "kdv": 600,
         "kategoriler": {"no": 3, "ad": "Market"}, "odeme_sekilleri": {"kod": "IGT_KK"},
         "projeler": {"ad": "Bolu Demiryolu"}, "personel": {"ad_soyad": "Halil Karakaş"}},
        {"tarih": f"{ay:%Y-%m}-03", "firma": "Opet", "fis_no": "A-1182",
         "aciklama": "Sondaj jeneratör yakıtı", "plaka_stok": "Mazot", "fis_tutari": 4200, "kdv": 700,
         "kategoriler": {"no": 23, "ad": "Sondaj Yakıt"}, "odeme_sekilleri": {"kod": "NAKIT_AVANS"},
         "projeler": {"ad": "Bolu Demiryolu"}, "personel": {"ad_soyad": "Mert Demir"}},
        {"tarih": f"{ay:%Y-%m}-07", "firma": "Köfteci Yusuf", "fis_no": "338",
         "aciklama": "Ekip öğle yemeği", "plaka_stok": None, "fis_tutari": 900, "kdv": 90,
         "kategoriler": {"no": 2, "ad": "Yemek"}, "odeme_sekilleri": {"kod": "PERSONEL_NAKIT"},
         "projeler": {"ad": "Bolu Demiryolu"}, "personel": {"ad_soyad": "Mert Demir"}},
        {"tarih": f"{ay:%Y-%m}-12", "firma": "Migros", "fis_no": "5521",
         "aciklama": "Şantiye sarf", "plaka_stok": None, "fis_tutari": 1250, "kdv": 250,
         "kategoriler": {"no": 8, "ad": "Sarf Malzeme"}, "odeme_sekilleri": {"kod": "IGT_KK"},
         "projeler": {"ad": "Geomembran İmalat"}, "personel": {"ad_soyad": "Ayşe Yıldız"}},
    ]
    a = [
        {"tarih": f"{ay:%Y-%m}-02", "tutar": 3000, "tur": "avans_verildi",
         "aciklama": "Nakit avans", "personel": {"ad_soyad": "Mert Demir"}},
        {"tarih": f"{ay:%Y-%m}-11", "tutar": 1500, "tur": "avans_verildi",
         "aciklama": "Nakit avans", "personel": {"ad_soyad": "Halil Karakaş"}},
    ]
    kategoriler = [
        {"no": 1, "ad": "Fatura"}, {"no": 2, "ad": "Yemek"}, {"no": 3, "ad": "Market"},
        {"no": 4, "ad": "Yakıt"}, {"no": 5, "ad": "Araç Bakım"}, {"no": 6, "ad": "Araç Yıkama"},
        {"no": 7, "ad": "Araç Kira"}, {"no": 8, "ad": "Sarf Malzeme"}, {"no": 9, "ad": "İş Kıyafeti"},
        {"no": 10, "ad": "Kırtasiye"}, {"no": 11, "ad": "Kargo"}, {"no": 12, "ad": "Ekipman"},
        {"no": 13, "ad": "Nakliye"}, {"no": 14, "ad": "Makine Tamir"}, {"no": 15, "ad": "Taşeron Hizmeti"},
        {"no": 16, "ad": "Kiralık Ekipman"}, {"no": 17, "ad": "Ulaşım"}, {"no": 18, "ad": "Konaklama"},
        {"no": 19, "ad": "Otopark"}, {"no": 20, "ad": "Alınan Emtia"}, {"no": 21, "ad": "Sağlık"},
        {"no": 22, "ad": "Temsili Ağırlama"}, {"no": 23, "ad": "Sondaj Yakıt"}, {"no": 24, "ad": "Sondaj Su"},
    ]
    return h, a, kategoriler


# ────────────────────────────────────────────────────────────────────
# Excel üretimi
# ────────────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BASLIK_FILL = PatternFill("solid", fgColor="1F3864")
BASLIK_FONT = Font(name=FONT_AD, bold=True, color="FFFFFF", size=10)
OZET_FILL = PatternFill("solid", fgColor="D9E1F2")
OZET_BASLIK_FILL = PatternFill("solid", fgColor="305496")
VURGU_FILL = PatternFill("solid", fgColor="FFF2CC")


def _g(d, *yol, default=None):
    """İç içe sözlükten güvenli okuma: _g(row,'kategoriler','ad')."""
    for k in yol:
        if not isinstance(d, dict):
            return default
        d = d.get(k)
    return d if d is not None else default


def _parse_tarih(s):
    if isinstance(s, (dt.date, dt.datetime)):
        return s if isinstance(s, dt.date) else s.date()
    return dt.date.fromisoformat(str(s)[:10])


def excel_uret(ay: dt.date, harcamalar, avanslar, kategoriler, cikti: str):
    wb = Workbook()
    ws = wb.active
    ws.title = TR_AYLAR[ay.month]

    # ── Ana tablo başlıkları (B1:K1), A = sıra no
    basliklar = ["", "TARİH", "FİRMA", "FATURA/FİŞ NO.", "AÇIKLAMA",
                 "PLAKA NO/STOK ADI", "H. ŞEKLİ", "FİŞ TUTARI", "KDV", "MATRAH", "H. YAPAN"]
    for i, b in enumerate(basliklar, start=1):
        c = ws.cell(row=1, column=i, value=b)
        c.font = BASLIK_FONT
        c.fill = BASLIK_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    # ── Veri satırları
    ilk_veri = 2
    r = ilk_veri
    for idx, h in enumerate(harcamalar, start=1):
        kod = _g(h, "odeme_sekilleri", "kod", default="")
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=_parse_tarih(h["tarih"])).number_format = TARIH_FORMAT
        ws.cell(row=r, column=3, value=_g(h, "firma"))
        ws.cell(row=r, column=4, value=_g(h, "fis_no"))
        ws.cell(row=r, column=5, value=_g(h, "aciklama"))
        ws.cell(row=r, column=6, value=_g(h, "plaka_stok"))
        ws.cell(row=r, column=7, value=ODEME_KISA.get(kod, kod))
        ws.cell(row=r, column=8, value=float(h.get("fis_tutari") or 0)).number_format = PARA_FORMAT
        ws.cell(row=r, column=9, value=float(h.get("kdv") or 0)).number_format = PARA_FORMAT
        # MATRAH = FİŞ TUTARI - KDV  (formül → DB ile birebir, canlı)
        ws.cell(row=r, column=10, value=f"=H{r}-I{r}").number_format = PARA_FORMAT
        ws.cell(row=r, column=11, value=_g(h, "personel", "ad_soyad"))
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    son_veri = r - 1 if harcamalar else ilk_veri
    # TOPLAM satırı (ana tablonun altı)
    toplam_satir = r + 1
    ws.cell(row=toplam_satir, column=7, value="TOPLAM").font = Font(name=FONT_AD, bold=True)
    if harcamalar:
        ws.cell(row=toplam_satir, column=8, value=f"=SUM(H{ilk_veri}:H{son_veri})")
        ws.cell(row=toplam_satir, column=9, value=f"=SUM(I{ilk_veri}:I{son_veri})")
        ws.cell(row=toplam_satir, column=10, value=f"=SUM(J{ilk_veri}:J{son_veri})")
    else:
        for col in (8, 9, 10):
            ws.cell(row=toplam_satir, column=col, value=0)
    for col in (8, 9, 10):
        c = ws.cell(row=toplam_satir, column=col)
        c.number_format = PARA_FORMAT
        c.font = Font(name=FONT_AD, bold=True)

    rng_H = f"$G${ilk_veri}:$G${son_veri}"
    rng_TUTAR = f"$H${ilk_veri}:$H${son_veri}"
    has_data = bool(harcamalar)

    # ── ÖZET TABLO (M:N) ─────────────────────────────────────────────
    def mn(row, m_val, n_val=None, m_bold=True, fill=None, n_fmt=PARA_FORMAT,
           comment=None):
        mc = ws.cell(row=row, column=13, value=m_val)
        mc.font = Font(name=FONT_AD, bold=m_bold, size=10)
        if fill:
            mc.fill = fill
        if n_val is not None:
            nc = ws.cell(row=row, column=14, value=n_val)
            nc.number_format = n_fmt
            nc.font = Font(name=FONT_AD, bold=m_bold, size=10)
            if fill:
                nc.fill = fill
        if comment:
            mc.comment = Comment(comment, "IGT Export")

    ws.merge_cells("M1:N1")
    h1 = ws.cell(row=1, column=13, value="ÖZET TABLO")
    h1.font = Font(name=FONT_AD, bold=True, color="FFFFFF", size=11)
    h1.fill = OZET_BASLIK_FILL
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=14).fill = OZET_BASLIK_FILL

    ws.cell(row=2, column=13, value="AY").font = Font(name=FONT_AD, bold=True)
    ac = ws.cell(row=2, column=14, value=ay.replace(day=1))
    ac.number_format = "MMMM YYYY"
    mn(3, "Harcamayı Yapan",
       n_val=", ".join(sorted({_g(h, "personel", "ad_soyad", default="") for h in harcamalar})) or "—",
       n_fmt="General")
    mn(4, "MATRAH", f"=SUM(J{ilk_veri}:J{son_veri})" if has_data else 0, fill=OZET_FILL)
    mn(5, "KDV", f"=SUM(I{ilk_veri}:I{son_veri})" if has_data else 0, fill=OZET_FILL)
    mn(6, "TOPLAM", "=N4+N5", fill=VURGU_FILL)

    # NAKİT/HAVALE/EFT İŞ AVANSLARI
    ws.merge_cells("M9:N9")
    a9 = ws.cell(row=9, column=13, value="NAKİT/HAVALE/EFT İŞ AVANSLARI")
    a9.font = Font(name=FONT_AD, bold=True, color="FFFFFF")
    a9.fill = OZET_BASLIK_FILL
    a9.alignment = Alignment(horizontal="center")
    ws.cell(row=9, column=14).fill = OZET_BASLIK_FILL
    ws.cell(row=10, column=13, value="Tarih").font = Font(name=FONT_AD, bold=True)
    ws.cell(row=10, column=14, value="Tutar (TL)").font = Font(name=FONT_AD, bold=True)

    avans_basla = 11
    ar = avans_basla
    avans_verildi = [a for a in avanslar if a.get("tur") == "avans_verildi"]
    for a in avans_verildi:
        ws.cell(row=ar, column=13, value=_parse_tarih(a["tarih"])).number_format = TARIH_FORMAT
        ws.cell(row=ar, column=14, value=float(a.get("tutar") or 0)).number_format = PARA_FORMAT
        ar += 1
    avans_son = ar - 1 if avans_verildi else avans_basla

    # Alt özet bloğu (sabit satırlar — orijinal formdaki düzen)
    mn(31, "TOPLAM NAKİT İŞ AVANSLARI",
       f"=SUM(N{avans_basla}:N{avans_son})" if avans_verildi else 0, fill=OZET_FILL)
    mn(32, ODEME_OZET_SATIR["NAKIT_AVANS"],
       f'=SUMIF({rng_H},"{ODEME_KISA["NAKIT_AVANS"]}",{rng_TUTAR})' if has_data else 0)
    mn(33, ODEME_OZET_SATIR["IGT_KK"],
       f'=SUMIF({rng_H},"{ODEME_KISA["IGT_KK"]}",{rng_TUTAR})' if has_data else 0)
    mn(34, ODEME_OZET_SATIR["PERSONEL_NAKIT"],
       f'=SUMIF({rng_H},"{ODEME_KISA["PERSONEL_NAKIT"]}",{rng_TUTAR})' if has_data else 0)
    mn(35, ODEME_OZET_SATIR["PERSONEL_KK"],
       f'=SUMIF({rng_H},"{ODEME_KISA["PERSONEL_KK"]}",{rng_TUTAR})' if has_data else 0)

    onceki_devir = sum(float(a.get("tutar") or 0)
                       for a in avanslar if a.get("tur") == "onceki_aydan_devir")
    mn(36, "ÖNCEKİ AYDAN DEVİR", onceki_devir if onceki_devir else None,
       comment="İş avansı / mutabakat modülü tamamlanınca otomatik dolacak (Faz 2).")
    mn(37, "PERSONEL ALACAK/BORÇ BAKİYESİ", None,
       comment="Mutabakat formülü iş avansı modülüyle netleşecek; şimdilik elle doldurulur.")

    # TAŞERON özet bloğu (yapısal — kaynak veri Faz 2'de bağlanacak)
    ws.merge_cells("M39:N39")
    t39 = ws.cell(row=39, column=13, value="TAŞERON")
    t39.font = Font(name=FONT_AD, bold=True, color="FFFFFF")
    t39.fill = OZET_BASLIK_FILL
    t39.alignment = Alignment(horizontal="center")
    ws.cell(row=39, column=14).fill = OZET_BASLIK_FILL
    mn(40, "İGT KREDİ KARTI", None)
    mn(41, "Nakit/Şahsi Kredi Kartı", None)
    mn(42, "TOPLAM", "=N40+N41", fill=VURGU_FILL)

    # ── MASRAF KALEMLERİ referans listesi (P:Q) ──────────────────────
    pq_h_no = ws.cell(row=1, column=16, value="NO")
    pq_h_ad = ws.cell(row=1, column=17, value="MASRAF KALEMLERİ")
    for c in (pq_h_no, pq_h_ad):
        c.font = BASLIK_FONT
        c.fill = BASLIK_FILL
        c.alignment = Alignment(horizontal="center")
    pr = 2
    for k in kategoriler:
        ws.cell(row=pr, column=16, value=k.get("no"))
        ws.cell(row=pr, column=17, value=f"*{k.get('ad')}")
        pr += 1

    # ── Sütun genişlikleri & görünüm
    genislik = {"A": 5, "B": 12, "C": 22, "D": 14, "E": 26, "F": 18, "G": 13,
                "H": 13, "I": 11, "J": 13, "K": 16, "L": 2,
                "M": 34, "N": 14, "O": 2, "P": 5, "Q": 22}
    for col, w in genislik.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    wb.save(cikti)
    return son_veri, len(harcamalar)


def main():
    p = argparse.ArgumentParser(description="IGT aylık masraf Excel dışa aktarımı")
    p.add_argument("--ay", required=True, help="YYYY-MM (örn. 2026-05)")
    p.add_argument("--cikti", default=None, help="Çıktı .xlsx yolu")
    p.add_argument("--mock", action="store_true", help="Canlı DB yerine örnek veri kullan")
    args = p.parse_args()

    try:
        yil, ayno = map(int, args.ay.split("-"))
        ay = dt.date(yil, ayno, 1)
    except Exception:
        sys.exit("HATA: --ay biçimi YYYY-MM olmalı (örn. 2026-05).")

    cikti = args.cikti or f"IGT-Masraf-{args.ay}.xlsx"
    harcamalar, avanslar, kategoriler = (mock_cek(ay) if args.mock else supabase_cek(ay))
    excel_uret(ay, harcamalar, avanslar, kategoriler, cikti)
    print(f"✓ {len(harcamalar)} harcama, {len(avanslar)} avans → {cikti}")


if __name__ == "__main__":
    main()
